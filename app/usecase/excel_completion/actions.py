import asyncio
import logging
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import List

import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.adapter.exception.app_exception import AppError
from app.core.constants import LOGGER_NAME
from app.core.eco_index.scraper import EcoindexScraper
from app.core.insight.google_insight import MobileInsight
from app.core.inspect_network.count_requests import InspectNetWork
from app.usecase.excel_completion.files_infos import (
    LIST_PAGE_NAME,
    SYNTHESE_PAGE_NAME,
    TEMPLATE_PATH,
    get_output_path,
)

logger = logging.getLogger(LOGGER_NAME)


def open_excel_file(fichier: Path | str) -> None:
    fichier = str(fichier)
    try:
        if sys.platform == "win32":
            os.startfile(fichier)
        elif sys.platform == "darwin":
            os.system(f"open {fichier}")
        else:
            os.system(f"xdg-open {fichier}")
    except Exception as e:
        raise AppError("Cannot open excel file") from e


def copy_sheet(
    source_sheet: Worksheet, target_workbook: Workbook, new_title: str
) -> None:
    target_sheet = target_workbook.create_sheet(title=new_title)

    for row in source_sheet.iter_rows():
        for cell in row:
            target_cell = target_sheet.cell(row=cell.row, column=cell.column)
            target_cell.value = cell.value
            if cell.has_style:
                target_cell.font = openpyxl.styles.Font(**dict(cell.font.__dict__))
                target_cell.border = openpyxl.styles.Border(
                    **dict(cell.border.__dict__)
                )
                target_cell.fill = openpyxl.styles.PatternFill(
                    **dict(cell.fill.__dict__)
                )
                target_cell.number_format = cell.number_format
                target_cell.protection = openpyxl.styles.Protection(
                    **dict(cell.protection.__dict__)
                )
                target_cell.alignment = openpyxl.styles.Alignment(
                    **dict(cell.alignment.__dict__)
                )

    for col, col_dim in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[col].width = col_dim.width
        target_sheet.column_dimensions[col].hidden = col_dim.hidden

    for row, row_dim in source_sheet.row_dimensions.items():
        target_sheet.row_dimensions[row].height = row_dim.height
        target_sheet.row_dimensions[row].hidden = row_dim.hidden


def create_excel_from_template(
    template_path: Path, output_path: str, urls: List[str]
) -> None:
    # Open the template
    template_wb: Workbook = openpyxl.load_workbook(template_path)

    # Create a new workbook
    new_wb: Workbook = openpyxl.Workbook()
    new_wb.remove(new_wb.active)

    # Copy the "Synthèse" sheet without modifying it
    synth_sheet = template_wb[SYNTHESE_PAGE_NAME]
    copy_sheet(synth_sheet, new_wb, SYNTHESE_PAGE_NAME)

    # Copy the "Liste" sheet without modifying it
    list_sheet = template_wb[LIST_PAGE_NAME]
    copy_sheet(list_sheet, new_wb, LIST_PAGE_NAME)

    # For each URL, create a new sheet based on "Page 1" from the template
    for i, url in enumerate(urls, start=1):
        logger.info("Analyse de la page %s ...", url)
        new_sheet_name: str = f"page {i}"
        copy_sheet(template_wb["page 1"], new_wb, new_sheet_name)

        insight = MobileInsight(url).get_result()
        logger.info("Insights google obtenus ...")

        eco_index = asyncio.run(EcoindexScraper(url=url).get_page_analysis())
        logger.info("Eco index obtenu ...")

        inspect = InspectNetWork(url=url).get_result()
        logger.info("Inspection du network completée ....")

        # url / date
        new_wb[new_sheet_name]["B3"] = url
        new_wb[new_sheet_name]["B4"] = datetime.now().strftime("%d/%m/%Y, %H:%M")

        # Green IT Analysis
        new_wb[new_sheet_name]["B12"] = eco_index.ges
        new_wb[new_sheet_name]["B13"] = f"{eco_index.size / 1000:.2f}"
        new_wb[new_sheet_name]["B14"] = eco_index.nodes
        new_wb[new_sheet_name]["B15"] = eco_index.requests

        # Lighthouse
        new_wb[new_sheet_name]["B18"] = insight.performance
        new_wb[new_sheet_name]["B19"] = f"{insight.first_contentful_paint / 1000:.2f}"
        new_wb[new_sheet_name]["C19"] = (
            f"Largest contentful paint : {insight.largest_contentful_paint / 1000:.2f} s"
        )
        new_wb[new_sheet_name]["B20"] = f"{insight.total_blocking_time / 1000:.2f}"
        new_wb[new_sheet_name]["B21"] = f"ok, {insight.speed_index} ms"

        # Réseau
        new_wb[new_sheet_name]["B24"] = inspect.total
        new_wb[new_sheet_name]["B25"] = inspect.js
        new_wb[new_sheet_name]["B26"] = inspect.css

        logger.info("Page %s pour l'url %s ajoutée", i, url)

    new_wb.save(output_path)


if __name__ == "__main__":
    # Usage example

    output_path = get_output_path()
    urls: List[str] = [
        "https://www.francetravail.fr/accueil",
        "https://www.alextraveylan.fr/fr",
        "https://it-wars.com/",
    ]

    create_excel_from_template(TEMPLATE_PATH, output_path, urls)
